import requests
from bs4 import BeautifulSoup
from urllib.parse import urljoin, urlparse
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ===============================
# BASIC HELPERS
# ===============================

def get_html(url):
    try:
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        return r.text
    except:
        return None

def same_domain(link, domain):
    return urlparse(link).netloc == domain

# ===============================
# DISCOVER REPOS UNDER PREFIX
# ===============================

def discover_repos(prefix):
    html = get_html(prefix)
    if not html:
        print("❌ Unable to load prefix")
        sys.exit()

    soup = BeautifulSoup(html, "html.parser")
    domain = urlparse(prefix).netloc
    repos = set()

    for a in soup.find_all("a", href=True):
        full = urljoin(prefix, a["href"])
        parsed = urlparse(full)

        if parsed.netloc == domain:
            parts = parsed.path.strip("/").split("/")
            if parts:
                repos.add(urljoin(prefix, parts[0] + "/"))

    return sorted(repos)

# ===============================
# PAGE ANALYSIS
# ===============================

def analyze_page(url, project_root):
    html = get_html(url)
    if not html:
        return None

    soup = BeautifulSoup(html, "html.parser")
    domain = urlparse(project_root).netloc

    internal, external = set(), set()

    for a in soup.find_all("a", href=True):
        link = urljoin(url, a["href"])
        if link.startswith("http"):
            if same_domain(link, domain):
                internal.add(link.split("#")[0])
            else:
                external.add(link)

    return {
        "url": url,
        "internal": sorted(internal),
        "external": sorted(external)
    }

# ===============================
# PROJECT ANALYSIS
# ===============================

def analyze_project(project_url):
    print(f"\n📁 Analyzing project: {project_url}")

    root = analyze_page(project_url, project_url)
    if not root:
        return []

    to_visit = set(root["internal"])
    visited = set()
    reports = []

    while to_visit:
        page = to_visit.pop()
        if page in visited or not page.startswith(project_url):
            continue

        visited.add(page)
        report = analyze_page(page, project_url)
        if report:
            reports.append(report)
            to_visit.update(report["internal"])

    return reports

# ===============================
# EXCEL EXPORT (STYLISH + DETAILED)
# ===============================

def export_excel(prefix, all_reports):
    wb = Workbook()

    # -------- SUMMARY SHEET --------
    summary = wb.active
    summary.title = "Summary"

    header_fill = PatternFill("solid", fgColor="CCE5FF")
    header_font = Font(bold=True)

    summary_headers = [
        "Project URL",
        "Total Pages",
        "Total Internal Links",
        "Total External Links"
    ]

    summary.append(summary_headers)
    for c in range(1, len(summary_headers) + 1):
        cell = summary.cell(row=1, column=c)
        cell.fill = header_fill
        cell.font = header_font

    for project, pages in all_reports.items():
        internal_count = sum(len(p["internal"]) for p in pages)
        external_count = sum(len(p["external"]) for p in pages)

        summary.append([
            project,
            len(pages),
            internal_count,
            external_count
        ])

    # Auto width
    for col in range(1, summary.max_column + 1):
        summary.column_dimensions[get_column_letter(col)].width = 35

    # -------- PROJECT SHEETS --------
    for project, pages in all_reports.items():
        sheet_name = project.rstrip("/").split("/")[-1][:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = [
            "Page URL",
            "Link Type",
            "Link URL"
        ]

        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font

        for page in pages:
            for link in page["internal"]:
                ws.append([page["url"], "Internal", link])

            for link in page["external"]:
                ws.append([page["url"], "External", link])

        # Auto width
        ws.column_dimensions["A"].width = 45
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 70

    filename = "Project_Analysis_Report.xlsx"
    wb.save(filename)
    print(f"\n📊 Excel report generated successfully: {filename}")

# ===============================
# MAIN FLOW
# ===============================

def main():
    prefix = input("Enter prefix URL: ").strip()

    print("\n🔍 Discovering REPO projects under prefix...")
    repos = discover_repos(prefix)

    if not repos:
        print("❌ No projects found")
        return

    print("\n📦 Projects found:")
    for i, r in enumerate(repos, 1):
        print(f"{i}. {r}")

    all_reports = {}

    for repo in repos:
        pages = analyze_project(repo)
        if pages:
            all_reports[repo] = pages

    while True:
        print("\nChoose an option:")
        print("1 : Export Detailed Excel Report")
        print("2 : Restart")
        print("3 : Exit")

        choice = input("> ").strip()

        if choice == "1":
            export_excel(prefix, all_reports)
        elif choice == "2":
            print("\n🔄 Restarting...\n")
            main()
            return
        elif choice == "3":
            print("👋 Exit")
            return
        else:
            print("❌ Invalid option")

# ===============================
# START
# ===============================

if __name__ == "__main__":
    main()